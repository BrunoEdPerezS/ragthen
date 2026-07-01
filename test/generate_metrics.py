"""
Genera metricas de calidad del sistema RAG para cada libreria y documento.
Accede a ChromaDB directamente (sin subprocess calls) para ser rapido.

Uso:
    python test/generate_metrics.py

Salida:
    metrics/ragthen_metrics.xlsx  — historial de metricas
    metrics/metrics_latest.json   — snapshot JSON de la corrida actual

Metricas por documento:
  - paginas, chunks, chars total, avg chars/chunk
  - secciones detectadas, FFFD rate, avg relevance
  - reader usado (cloud si tiene secciones, local si no)
"""
import json
import subprocess
from datetime import date
from pathlib import Path

import chromadb

LIBRARIES_DIR = Path.home() / ".ragthen" / "libraries"
REPO_ROOT = Path(__file__).resolve().parent.parent
METRICS_DIR = REPO_ROOT / "metrics"
EXCEL_PATH = METRICS_DIR / "ragthen_metrics.xlsx"


def get_git_commit() -> str:
    try:
        r = subprocess.run(
            ["git", "log", "--oneline", "-1"],
            capture_output=True, text=True, timeout=10,
            cwd=REPO_ROOT,
        )
        return r.stdout.strip().split()[0] if r.stdout else "unknown"
    except Exception:
        return "unknown"


def count_pdf_pages(filepath: Path) -> int:
    try:
        import fitz
        doc = fitz.open(str(filepath))
        n = len(doc)
        doc.close()
        return n
    except Exception:
        return 0


def get_library_docs(lib_path: Path) -> list[dict]:
    """Lista documentos fisicos en la carpeta de la libreria."""
    docs = []
    for f in sorted(lib_path.iterdir()):
        if f.is_file() and f.suffix.lower() in (".pdf", ".epub", ".txt", ".md"):
            pages = count_pdf_pages(f) if f.suffix.lower() == ".pdf" else 0
            docs.append({"name": f.name, "path": f, "pages": pages})
    return docs


def load_chroma_data(index_dir: Path) -> dict | None:
    """Carga todos los chunks de ChromaDB y retorna metadata agregada."""
    if not index_dir.exists():
        return None
    try:
        db = chromadb.PersistentClient(path=str(index_dir))
        col = db.get_collection("ragthen")
        count = col.count()
        if count == 0:
            return None

        batch_size = 1000
        all_chunks = []
        for offset in range(0, count, batch_size):
            batch = col.get(
                limit=min(batch_size, count - offset),
                offset=offset,
                include=["documents", "metadatas"],
            )
            for doc_text, meta in zip(batch["documents"], batch["metadatas"]):
                all_chunks.append({"text": doc_text or "", "metadata": meta})

        return {"total_chunks": count, "chunks": all_chunks}
    except Exception as e:
        print(f"    Error reading ChromaDB: {e}")
        return None


def compute_doc_metrics(chunks: list[dict], doc_name: str, pages: int) -> dict:
    """Calcula metricas para chunks que pertenecen a un documento."""
    doc_chunks = [c for c in chunks if c["metadata"].get("source", "").endswith(doc_name)]

    total_chars = sum(len(c["text"]) for c in doc_chunks)
    fffd_count = sum(c["text"].count("\ufffd") for c in doc_chunks)
    sections = set()
    for c in doc_chunks:
        sec = c["metadata"].get("section", "")
        if sec:
            sections.add(sec)

    has_sections = len(sections) > 0

    return {
        "documento": doc_name,
        "paginas": pages,
        "chunks": len(doc_chunks),
        "chars_total": total_chars,
        "avg_chars_chunk": round(total_chars / max(len(doc_chunks), 1), 1),
        "secciones_detectadas": len(sections),
        "fffd_rate": round(fffd_count / max(total_chars, 1) * 1000, 4),
        "reader": "cloud" if has_sections else "local",
        "notas": "",
    }


def compute_library_avg_relevance(index_dir: Path, num_samples: int = 100) -> float:
    """Calcula avg relevance como la distancia promedio entre chunks."""
    try:
        db = chromadb.PersistentClient(path=str(index_dir))
        col = db.get_collection("ragthen")
        count = col.count()
        if count < 2:
            return 0.0

        sample = col.get(limit=min(num_samples, count), include=["embeddings"])
        embs = sample["embeddings"]
        if not embs or len(embs) < 2:
            return 0.0

        import numpy as np
        from sklearn.metrics.pairwise import cosine_similarity

        mat = np.array(embs)
        sim = cosine_similarity(mat)
        n = sim.shape[0]
        scores = (sim.sum() - n) / (n * (n - 1))
        return round(max(0, float(scores)), 4)
    except Exception:
        return 0.0


def main():
    commit = get_git_commit()
    today = date.today().isoformat()
    METRICS_DIR.mkdir(parents=True, exist_ok=True)

    libs = sorted(d.name for d in LIBRARIES_DIR.iterdir()
                  if d.is_dir() and not d.name.startswith("."))

    all_rows = []

    for lib_name in libs:
        print(f"\n{'='*60}")
        print(f"Library: {lib_name}")
        print(f"{'='*60}")

        lib_path = LIBRARIES_DIR / lib_name
        index_dir = lib_path / ".chroma"

        chroma_data = load_chroma_data(index_dir)
        if chroma_data is None:
            print(f"  (not indexed or empty)")
            continue

        docs_fs = get_library_docs(lib_path)
        all_chunks = chroma_data["chunks"]

        print(f"  Total chunks: {chroma_data['total_chunks']}")
        print(f"  Documentos fisicos: {len(docs_fs)}")

        avg_rel = compute_library_avg_relevance(index_dir)
        print(f"  Avg relevance: {avg_rel}")

        for doc in docs_fs:
            metrics = compute_doc_metrics(all_chunks, doc["name"], doc["pages"])
            metrics["libreria"] = lib_name
            metrics["fecha"] = today
            metrics["commit"] = commit
            metrics["avg_relevance"] = avg_rel
            all_rows.append(metrics)

            status = f"  {doc['name']}: {metrics['chunks']} chunks, "
            status += f"{metrics['chars_total']} chars, "
            status += f"{metrics['secciones_detectadas']} secciones, "
            status += f"FFFD={metrics['fffd_rate']}"
            print(status)

    print(f"\n{'='*60}")
    print(f"Generando Excel y JSON...")

    json_path = METRICS_DIR / "metrics_latest.json"
    json_path.write_text(json.dumps(all_rows, ensure_ascii=False, indent=2),
                         encoding="utf-8")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = f"RAG Metrics {today}"

        headers = ["fecha", "commit", "libreria", "documento", "paginas",
                   "chunks", "chars_total", "avg_chars_chunk",
                   "secciones_detectadas", "fffd_rate", "avg_relevance",
                   "reader", "notas"]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                                  fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(all_rows, 2):
            for col_idx, h in enumerate(headers, 1):
                val = row_data.get(h, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center")

        col_widths = [12, 10, 15, 50, 10, 10, 15, 15, 20, 12, 15, 10, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = w

        wb.save(str(EXCEL_PATH))
        print(f"Excel guardado: {EXCEL_PATH}")
    except ImportError:
        print("openpyxl no instalado. pip install openpyxl")

    print(f"\nResumen: {len(all_rows)} documentos en {len(libs)} librerias, "
          f"commit {commit}")


if __name__ == "__main__":
    main()
